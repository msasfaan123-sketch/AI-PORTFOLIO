import random
import re
import sys
import textwrap
import json
import math
import os
import urllib.error
import urllib.request
from datetime import datetime
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path

import pandas as pd

try:
    from flask import Flask, jsonify, request
    FLASK_AVAILABLE = True
except ImportError:
    FLASK_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from pypdf import PdfReader
    PDF_READER_AVAILABLE = True
except ImportError:
    PDF_READER_AVAILABLE = False

try:
    import pyttsx3
    TTS_AVAILABLE = True
except ImportError:
    TTS_AVAILABLE = False

try:
    import speech_recognition as sr
    SPEECH_RECOGNITION_AVAILABLE = True
except ImportError:
    SPEECH_RECOGNITION_AVAILABLE = False

try:
    from rapidfuzz import fuzz
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False

    class fuzz:
        @staticmethod
        def ratio(left, right):
            return SequenceMatcher(None, left, right).ratio() * 100

        @staticmethod
        def partial_ratio(left, right):
            if not left or not right:
                return 0

            shorter, longer = sorted((left, right), key=len)

            if shorter in longer:
                return 100

            best = 0
            matcher = SequenceMatcher(None, shorter, longer)

            for block in matcher.get_matching_blocks():
                start = max(block.b - block.a, 0)
                candidate = longer[start:start + len(shorter)]

                if not candidate:
                    continue

                score = fuzz.ratio(shorter, candidate)

                if score > best:
                    best = score

                    if best >= 99:
                        return best

            return best

        @staticmethod
        def token_sort_ratio(left, right):
            return fuzz.ratio(
                " ".join(sorted(left.split())),
                " ".join(sorted(right.split())),
            )

        @staticmethod
        def token_set_ratio(left, right):
            left_words = set(left.split())
            right_words = set(right.split())

            if not left_words or not right_words:
                return 0

            common = left_words & right_words

            if common and (common == left_words or common == right_words):
                return 100

            combined_left = " ".join(sorted(common | (left_words - right_words)))
            combined_right = " ".join(sorted(common | (right_words - left_words)))

            return fuzz.ratio(combined_left, combined_right)


# =========================================================
# CONFIG
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent.parent


def load_local_env():
    env_file = PROJECT_DIR / ".env"

    if not env_file.exists():
        return

    for line in env_file.read_text(encoding="utf-8").splitlines():
        line = line.strip()

        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("\"'")

        if key and key not in os.environ:
            os.environ[key] = value


load_local_env()

EXCEL_FILE = BASE_DIR / "Asfaan_Interview_Questions_v3.xlsx"
RESUME_FILE = BASE_DIR / "MS.ASFAAN_RESUME.pdf"
MISSED_QUESTIONS_FILE = BASE_DIR / "missed_questions.txt"
TRANSCRIPT_DIR = BASE_DIR / "transcripts"
KNOWLEDGE_BASE_FILE = BASE_DIR / "knowledge_base.json"
OPENROUTER_CHAT_URL = "https://openrouter.ai/api/v1/chat/completions"
OPENROUTER_MODEL = os.environ.get("OPENROUTER_MODEL", "openrouter/free")

FUZZY_THRESHOLD = 40
DEBUG_SCORE = False
WRAP_WIDTH = 110

EXIT_COMMANDS = {
    "exit",
    "quit",
    "bye",
    "close",
    "stop",
}

SHORT_FOLLOWUPS = (
    "what about",
    "which one",
    "how",
    "why",
    "when",
    "where",
    "did you",
    "tell more",
    "continue",
    "what tech",
    "what database",
    "which stack",
)

DIRECT_INTENTS = (
    (
        (
            "who are you",
            "who you",
            "about you",
            "introduce yourself",
            "your introduction",
            "tell me about yourself",
        ),
        ("tell me about yourself",),
    ),
    (
        (
            "hire you",
            "why hire",
            "why should we hire you",
            "why choose you",
            "why you",
        ),
        ("why should we hire you",),
    ),
    (
        (
            "relocation",
            "move city",
            "open to move",
            "open to relocation",
            "are you open to relocation",
        ),
        ("open to relocation",),
    ),
    (
        (
            "salary",
            "expected salary",
            "salary expectation",
            "ctc",
            "package",
        ),
        ("salary", "expecting"),
    ),
    (
        (
            "strength",
            "strengths",
            "good at",
            "best skill",
        ),
        ("greatest strengths",),
    ),
    (
        (
            "weakness",
            "weaknesses",
            "bad at",
        ),
        ("biggest weakness",),
    ),
    (
        (
            "skills",
            "technical skills",
            "tech skills",
            "technology skills",
        ),
        ("technical skills",),
    ),
    (
        (
            "education",
            "study",
            "degree",
            "college",
        ),
        ("field of study",),
    ),
    (
        (
            "project",
            "projects",
            "your project",
            "portfolio project",
        ),
        ("project",),
    ),
)

GREETING_PATTERNS = {
    "hi",
    "hello",
    "hey",
    "good morning",
    "good afternoon",
    "good evening",
}

HELP_COMMANDS = {
    "help",
    "menu",
    "topics",
    "options",
}

ANSWER_MODES = {
    "normal",
    "short",
    "interview",
    "friendly",
    "professional",
    "recruiter",
}

MODE_ALIASES = {
    "brief": "short",
    "formal": "professional",
    "hr": "interview",
}

PRACTICE_STOP_COMMANDS = {
    "stop practice",
    "end practice",
    "quit practice",
}

TOKEN_STOPWORDS = {
    "about",
    "are",
    "ask",
    "can",
    "did",
    "for",
    "how",
    "me",
    "practice",
    "question",
    "questions",
    "show",
    "tell",
    "the",
    "was",
    "what",
    "when",
    "where",
    "which",
    "who",
    "why",
    "you",
    "your",
}

CATEGORY_ALIASES = {
    "hr": "hr round",
    "human resource": "hr round",
    "behavior": "general behavioral",
    "behaviour": "general behavioral",
    "behavioral": "general behavioral",
    "technical": "skills based",
    "tech": "skills based",
    "skill": "skills based",
    "skills": "skills based",
    "project": "projects intern",
    "projects": "projects intern",
    "intern": "projects intern",
    "internship": "projects intern",
    "personal": "etiquette personal",
}

TEXT_CLEANER = re.compile(r"[^a-z0-9\s]")
SPACE_CLEANER = re.compile(r"\s+")


# =========================================================
# MEMORY SYSTEM
# =========================================================

memory = {
    "last_topic": None,
    "conversation_depth": 0,
    "allowed_followups": (),
    "last_response": None,
    "last_row": None,
    "last_match": None,
    "answer_mode": "normal",
    "speak": False,
    "practice": None,
    "mock": None,
    "trainer": None,
    "transcript": [],
    "show_source": False,
}

web_state = None
last_openrouter_error = None
last_openrouter_selected_model = None


# =========================================================
# COMMON WORD NORMALIZATION
# =========================================================

COMMON_REPLACEMENTS = {
    "2": "to",
    "4": "for",
    "8": "ate",
    "cn": "can",
    "iam": "i am",
    "im": "i am",
    "plz": "please",
    "pls": "please",
    "r": "are",
    "relocatn": "relocation",
    "ur": "your",
    "u": "you",
    "wbu": "what about you",
    "wat": "what",
    "wht": "what",
    "abt": "about",
    "proj": "project",
    "projs": "projects",
    "db": "database",
    "wrk": "work",
    "wrks": "works",
    "intro": "introduction",
    "dev": "developer",
    "yr": "year",
    "yrs": "years",
}


# =========================================================
# TEXT CLEANER
# =========================================================

@lru_cache(maxsize=8192)
def clean_text(text):
    """
    Clean and normalize text. Cached because the same workbook text is reused
    across every chat turn.
    """

    if pd.isna(text):
        return ""

    text = str(text).lower()
    text = TEXT_CLEANER.sub(" ", text)
    text = SPACE_CLEANER.sub(" ", text).strip()

    if not text:
        return ""

    return " ".join(
        COMMON_REPLACEMENTS.get(word, word)
        for word in text.split()
    )


def clean_display_text(text):
    return SPACE_CLEANER.sub(" ", str(text).strip())


def print_bot(message):
    wrapped = textwrap.fill(clean_display_text(message), width=WRAP_WIDTH)
    print(f"\nBot: {wrapped}\n")

    if memory["speak"]:
        speak_text(message)


def finish_response(response, should_exit=False):
    memory["transcript"].append(
        {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "bot": clean_display_text(response),
        }
    )
    return response, should_exit


def openrouter_is_configured():
    return bool(os.environ.get("OPENROUTER_API_KEY"))


def build_llm_payload(user_input, local_answer, model):
    prompt = (
        "You are Mohamed Sathak Asfaan's portfolio assistant. "
        "Answer only using the provided portfolio context. "
        "Keep the answer concise, friendly, and recruiter-ready. "
        "If the context is not enough, say that the portfolio data does not include that detail."
    )
    return {
        "model": model,
        "messages": [
            {"role": "system", "content": prompt},
            {
                "role": "user",
                "content": (
                    f"Question: {clean_display_text(user_input)}\n\n"
                    f"Portfolio context: {clean_display_text(local_answer)}"
                ),
            },
        ],
        "temperature": 0.35,
        "max_tokens": 260,
    }


def post_chat_completion(url, api_key, payload, extra_headers=None):
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    if extra_headers:
        headers.update(extra_headers)

    request_data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(url, data=request_data, headers=headers, method="POST")

    with urllib.request.urlopen(request, timeout=12) as response:
        data = json.loads(response.read().decode("utf-8"))

    choices = data.get("choices", [])

    if not choices:
        return None, "provider returned no choices.", data.get("model")

    message = choices[0].get("message", {})
    content = clean_display_text(message.get("content", ""))

    if not content or content.lower() in {"none", "null", "n/a"}:
        return None, "provider returned an empty message.", data.get("model")

    return content, None, data.get("model")


def openrouter_generate_reply(user_input, local_answer):
    global last_openrouter_error, last_openrouter_selected_model
    last_openrouter_error = None
    last_openrouter_selected_model = None
    api_key = os.environ.get("OPENROUTER_API_KEY")

    if not api_key or not local_answer:
        last_openrouter_error = "OPENROUTER_API_KEY is not configured."
        return None

    payload = build_llm_payload(user_input, local_answer, OPENROUTER_MODEL)

    try:
        content, error, selected_model = post_chat_completion(
            OPENROUTER_CHAT_URL,
            api_key,
            payload,
            {
                "HTTP-Referer": os.environ.get("APP_URL", "http://localhost:8080"),
                "X-OpenRouter-Title": "Asfaan Portfolio Chatbot",
            },
        )
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8")
            error_data = json.loads(body)
            message = error_data.get("error", {}).get("message", body)
        except Exception:
            message = str(e)
        last_openrouter_error = f"OpenRouter HTTP {e.code}: {message}"
        return None
    except urllib.error.URLError as e:
        last_openrouter_error = f"OpenRouter network error: {e.reason}"
        return None
    except TimeoutError:
        last_openrouter_error = "OpenRouter request timed out."
        return None
    except (json.JSONDecodeError, KeyError) as e:
        last_openrouter_error = f"OpenRouter response parse error: {e}"
        return None

    last_openrouter_error = error
    last_openrouter_selected_model = selected_model
    return content


def remember_user_message(user_input):
    memory["transcript"].append(
        {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "user": clean_display_text(user_input),
        }
    )


def speak_text(message):
    if not TTS_AVAILABLE:
        return

    try:
        engine = pyttsx3.init()
        engine.say(clean_display_text(message))
        engine.runAndWait()
    except Exception:
        memory["speak"] = False


def listen_for_voice():
    if not SPEECH_RECOGNITION_AVAILABLE:
        return None, "Voice input needs SpeechRecognition installed."

    try:
        recognizer = sr.Recognizer()

        with sr.Microphone() as source:
            recognizer.adjust_for_ambient_noise(source, duration=0.4)
            audio = recognizer.listen(source, timeout=5, phrase_time_limit=10)

        return recognizer.recognize_google(audio), None
    except Exception as e:
        return None, f"I could not capture voice input: {e}"


# =========================================================
# LOAD AND PREPARE DATASET
# =========================================================

def load_dataset(file_path):
    """
    Load Excel dataset.
    """

    df = pd.read_excel(file_path)
    df.columns = [str(col).strip() for col in df.columns]
    return df


def load_knowledge_base(file_path):
    if not file_path.exists():
        return {}

    with file_path.open("r", encoding="utf-8") as file:
        return json.load(file)


def load_resume_text(file_path):
    if not PDF_READER_AVAILABLE or not file_path.exists():
        return ""

    try:
        reader = PdfReader(str(file_path))
        pages = [page.extract_text() or "" for page in reader.pages]
        return clean_display_text(" ".join(pages))
    except Exception:
        return ""


def prepare_resume_index(resume_text):
    chunks = []
    token_lookup = {}

    if not resume_text:
        return {"chunks": (), "token_lookup": {}}

    parts = re.split(r"(?<=[.!?])\s+|\n+", resume_text)

    for part in parts:
        part = clean_display_text(part)

        if len(part) < 30:
            continue

        cleaned = clean_text(part)
        tokens = frozenset(query_tokens(cleaned))

        chunk = {
            "text": part,
            "cleaned": cleaned,
            "tokens": tokens,
        }
        chunks.append(chunk)

        for token in tokens:
            token_lookup.setdefault(token, []).append(chunk)

    return {
        "chunks": tuple(chunks),
        "token_lookup": {
            token: tuple(token_chunks)
            for token, token_chunks in token_lookup.items()
        },
    }


def extract_contact_info(resume_text):
    email_match = re.search(r"[\w.+-]+@[\w.-]+\.\w+", resume_text)
    phone_match = re.search(r"(?:\+?\d[\d\s-]{8,}\d)", resume_text)
    email = email_match.group(0) if email_match else None

    if email and email.lower().startswith("envelope"):
        email = email[8:]

    return {
        "email": email,
        "phone": clean_display_text(phone_match.group(0)) if phone_match else None,
        "linkedin": "asfaan" if "linkedin" in resume_text.lower() else None,
        "github": "asfaan-dev" if "github" in resume_text.lower() else None,
    }


def get_matching_columns(columns, *needles):
    """
    Pick workbook columns by name once, instead of checking every cell on every
    user message.
    """

    return [
        col
        for col in columns
        if any(needle in str(col).lower() for needle in needles)
    ]


def collect_clean_values(row, columns):
    values = []

    for col in columns:
        value = row[col]

        if pd.notna(value):
            cleaned = clean_text(value)

            if cleaned:
                values.append(cleaned)

    return tuple(dict.fromkeys(values))


def collect_responses(row, columns):
    responses = []

    for col in columns:
        value = row[col]

        if pd.notna(value):
            value = str(value).strip()

            if value:
                responses.append(value)

    if not responses and "Question" in row.index and pd.notna(row["Question"]):
        responses.append(f"Ah yes, {row['Question']}")

    return tuple(responses)


def prepare_dataset(df):
    """
    Convert the workbook into a fast search index. Each entry stores pre-cleaned
    keywords, responses, follow-ups, and question text.
    """

    keyword_columns = get_matching_columns(
        df.columns,
        "question",
        "casual",
        "keyword",
    )
    response_columns = get_matching_columns(df.columns, "response", "answer")
    followup_columns = get_matching_columns(df.columns, "follow")

    rows = []
    followup_lookup = {}
    token_lookup = {}

    for _, row in df.iterrows():
        keywords = collect_clean_values(row, keyword_columns)
        responses = collect_responses(row, response_columns)
        cleaned_responses = tuple(clean_text(response) for response in responses)
        followups = collect_clean_values(row, followup_columns)
        question = clean_text(row.get("Question", ""))
        category = clean_text(row.get("Category", ""))
        tokens = frozenset(
            token
            for text in keywords + cleaned_responses + followups + (category,)
            for token in text.split()
            if len(token) > 2 and token not in TOKEN_STOPWORDS
        )

        prepared = {
            "question": question,
            "category": category,
            "keywords": keywords,
            "tokens": tokens,
            "responses": responses,
            "cleaned_responses": cleaned_responses,
            "followups": followups,
        }

        rows.append(prepared)

        for followup in followups:
            followup_lookup.setdefault(followup, prepared)

        for token in tokens:
            token_lookup.setdefault(token, []).append(prepared)

    return {
        "rows": tuple(rows),
        "followup_lookup": followup_lookup,
        "token_lookup": {
            token: tuple(token_rows)
            for token, token_rows in token_lookup.items()
        },
        "topics": tuple(sorted({row["category"] for row in rows if row["category"]})),
    }


# =========================================================
# MATCHING ENGINE
# =========================================================

def query_tokens(text):
    tokens = set()

    for token in text.split():
        if len(token) <= 2 or token in TOKEN_STOPWORDS:
            continue

        tokens.add(token)

        if token.endswith("s") and len(token) > 4:
            tokens.add(token[:-1])

    return tokens


def find_resume_match(user_input, resume_index):
    cleaned_input = clean_text(user_input)
    tokens = query_tokens(cleaned_input)

    if not tokens or not resume_index["chunks"]:
        return None, 0

    seen = set()
    candidates = []

    for token in tokens:
        for chunk in resume_index["token_lookup"].get(token, ()):
            chunk_id = id(chunk)

            if chunk_id not in seen:
                seen.add(chunk_id)
                candidates.append(chunk)

    if not candidates:
        return None, 0

    best_chunk = None
    best_score = 0

    for chunk in candidates:
        overlap = len(tokens & chunk["tokens"]) / max(len(tokens), 1) * 100
        fuzzy_score = fuzz.token_set_ratio(cleaned_input, chunk["cleaned"])
        score = overlap * 0.65 + fuzzy_score * 0.35

        if score > best_score:
            best_score = score
            best_chunk = chunk

    if best_score < 52:
        return None, best_score

    return best_chunk, best_score


def detect_project(cleaned_input, knowledge_base):
    projects = knowledge_base.get("projects", {})

    for project_id, project in projects.items():
        aliases = [project_id, project.get("name", ""), *project.get("aliases", [])]

        for alias in aliases:
            cleaned_alias = clean_text(alias)

            if cleaned_alias and cleaned_alias in cleaned_input:
                return project_id, project

    return None, None


def classify_intent(cleaned_input, knowledge_base):
    tokens = query_tokens(cleaned_input)
    project_id, project = detect_project(cleaned_input, knowledge_base)

    if project:
        if tokens & {"stack", "technology", "technologies", "tools", "used"}:
            return "project_tech", project_id, project

        if tokens & {"role", "contribution", "responsibility", "responsibilities"}:
            return "project_role", project_id, project

        if tokens & {"impact", "result", "results", "outcome", "benefit"}:
            return "project_impact", project_id, project

        if tokens & {"challenge", "challenges", "problem", "difficult"}:
            return "project_challenge", project_id, project

        if tokens & {"feature", "features", "work", "working", "workflow", "search", "database", "scale"}:
            return "project_detail", project_id, project

        return "project_summary", project_id, project

    if tokens & {"skill", "skills", "stack", "technology", "technologies"}:
        return "skills", None, None

    if tokens & {"education", "degree", "college", "study", "qualification"}:
        return "education", None, None

    if tokens & {"experience", "internship", "work"}:
        return "experience", None, None

    if tokens & {"achievement", "achievements", "impact", "result", "results"}:
        return "achievements", None, None

    if tokens & {"summary", "profile", "overview"}:
        return "profile", None, None

    return "unknown", None, None


def project_template_answer(intent, project):
    if intent == "project_tech":
        return f"{project['name']} used: {', '.join(project.get('tech_stack', []))}."

    if intent == "project_role":
        return f"Asfaan's role in {project['name']}: {project.get('role', project.get('summary', ''))}"

    if intent == "project_impact":
        return f"Impact of {project['name']}: {project.get('impact', project.get('summary', ''))}"

    if intent == "project_challenge":
        return f"Key challenges in {project['name']}: {', '.join(project.get('challenges', []))}."

    if intent == "project_detail":
        followups = project.get("followups", {})
        detail = " ".join(followups.values()) if followups else project.get("summary", "")
        features = ", ".join(project.get("features", []))
        return f"{project['name']} details: {detail} Main features: {features}."

    return (
        f"{project['name']}: {project.get('summary', '')} "
        f"Problem: {project.get('problem', '')} "
        f"Impact: {project.get('impact', '')}"
    )


def knowledge_answer(cleaned_input, state):
    knowledge_base = state.get("knowledge_base", {})

    if not knowledge_base:
        return None, None

    intent, project_id, project = classify_intent(cleaned_input, knowledge_base)

    if project:
        return project_template_answer(intent, project), f"knowledge base:{project_id}"

    profile = knowledge_base.get("profile", {})
    skills = knowledge_base.get("skills", {})

    if intent == "skills":
        return skills.get("summary"), "knowledge base:skills"

    if intent == "education":
        return profile.get("education"), "knowledge base:education"

    if intent == "experience":
        return profile.get("experience"), "knowledge base:experience"

    if intent == "achievements":
        achievements = knowledge_base.get("achievements", [])
        return "Key achievements: " + " | ".join(achievements), "knowledge base:achievements"

    if intent == "profile":
        return profile.get("summary"), "knowledge base:profile"

    return None, None


def build_similarity_documents(state):
    documents = []

    knowledge_base = state.get("knowledge_base", {})

    for project_id, project in knowledge_base.get("projects", {}).items():
        text = " ".join(
            [
                project.get("name", ""),
                project.get("summary", ""),
                project.get("problem", ""),
                project.get("role", ""),
                project.get("impact", ""),
                " ".join(project.get("tech_stack", [])),
                " ".join(project.get("features", [])),
                " ".join(project.get("challenges", [])),
                " ".join(project.get("aliases", [])),
            ]
        )
        documents.append(
            {
                "source": f"knowledge base:{project_id}",
                "text": text,
                "answer": project_template_answer("project_summary", project),
                "tokens": query_tokens(clean_text(text)),
            }
        )

    for row in state["index"]["rows"]:
        text = " ".join((row["question"], row["category"], *row["keywords"]))
        answer = row["responses"][0] if row["responses"] else ""
        documents.append(
            {
                "source": "excel",
                "text": text,
                "answer": answer,
                "tokens": query_tokens(text),
            }
        )

    for chunk in state["resume_index"]["chunks"]:
        documents.append(
            {
                "source": "resume",
                "text": chunk["cleaned"],
                "answer": "From the resume: " + chunk["text"],
                "tokens": chunk["tokens"],
            }
        )

    return tuple(documents)


def find_similarity_answer(cleaned_input, state):
    tokens = query_tokens(cleaned_input)

    if not tokens:
        return None, 0, None

    best_doc = None
    best_score = 0

    for doc in state.get("similarity_docs", ()):
        doc_tokens = doc["tokens"]

        if not doc_tokens:
            continue

        overlap = len(tokens & doc_tokens)
        cosine = overlap / math.sqrt(len(tokens) * len(doc_tokens))
        fuzzy_score = fuzz.token_set_ratio(cleaned_input, doc["text"]) / 100
        score = cosine * 0.7 + fuzzy_score * 0.3

        if score > best_score:
            best_score = score
            best_doc = doc

    if not best_doc or best_score < 0.34:
        return None, best_score, None

    return best_doc["answer"], best_score, best_doc["source"]


def validate_answer(cleaned_input, response, source):
    tokens = query_tokens(cleaned_input)

    if not tokens:
        return True

    response_tokens = query_tokens(clean_text(response))
    overlap = tokens & response_tokens

    if source and source.startswith("knowledge base"):
        return True

    if len(tokens) == 1:
        return bool(overlap)

    return len(overlap) >= 1


def find_row_by_targets(index, targets):
    target_tokens = tuple(query_tokens(" ".join(targets)))

    for row in index["rows"]:
        if all(target in row["question"] for target in targets):
            return row

        if target_tokens and all(token in row["question"].split() for token in target_tokens):
            return row

    for row in index["rows"]:
        keyword_text = " ".join(row["keywords"])

        if all(target in keyword_text for target in targets):
            return row

        if target_tokens and all(token in row["tokens"] for token in target_tokens):
            return row

    return None


def direct_intent_match(cleaned_input, index):
    for phrases, targets in DIRECT_INTENTS:
        for phrase in phrases:
            cleaned_phrase = clean_text(phrase)

            if cleaned_input == cleaned_phrase:
                return find_row_by_targets(index, targets), 100

            if cleaned_phrase in cleaned_input or cleaned_input in cleaned_phrase:
                return find_row_by_targets(index, targets), 96

            if fuzz.token_set_ratio(cleaned_input, cleaned_phrase) >= 92:
                return find_row_by_targets(index, targets), 94

    return None, 0


def best_keyword_score(user_input, keywords):
    best_score = 0

    for keyword in keywords:
        if user_input == keyword:
            return 100

        if user_input in keyword or keyword in user_input:
            best_score = max(best_score, 92)
            continue

        score = max(
            fuzz.partial_ratio(user_input, keyword),
            fuzz.token_sort_ratio(user_input, keyword),
            fuzz.token_set_ratio(user_input, keyword),
        )

        if score > best_score:
            best_score = score

    return best_score


def best_response_score(user_input, cleaned_responses):
    best_score = 0

    for response in cleaned_responses:
        if not response:
            continue

        if user_input == response:
            return 100

        if user_input in response or response in user_input:
            best_score = max(best_score, 85)
            continue

        if RAPIDFUZZ_AVAILABLE:
            score = fuzz.partial_ratio(user_input, response)
        else:
            score = fuzz.token_set_ratio(user_input, response)

        if score > best_score:
            best_score = score

    return best_score


def calculate_score(user_input, prepared_row):
    """
    Weighted fuzzy scoring against already-prepared row data.
    """

    keyword_score = best_keyword_score(user_input, prepared_row["keywords"])
    response_score = best_response_score(
        user_input,
        prepared_row["cleaned_responses"],
    )

    return keyword_score * 0.85 + response_score * 0.15


def dynamic_threshold(cleaned_input):
    token_count = len(query_tokens(cleaned_input))

    if token_count <= 2:
        return 68

    if token_count <= 4:
        return 54

    return FUZZY_THRESHOLD


def is_confident_match(cleaned_input, best_score, second_score):
    tokens = query_tokens(cleaned_input)

    if len(tokens) == 1 and best_score < 82:
        return False

    if best_score < dynamic_threshold(cleaned_input):
        return False

    if best_score >= 80:
        return True

    if second_score and best_score - second_score < 5:
        return False

    return True


# =========================================================
# CONTEXT BOOSTER
# =========================================================

def apply_context_boost(score, prepared_row, user_input):
    """
    Boost score using memory context.
    """

    last_topic = memory["last_topic"]

    if last_topic and last_topic in prepared_row["question"]:
        score += 10

    if any(phrase in user_input for phrase in SHORT_FOLLOWUPS):
        score += 8

    return score


# =========================================================
# FIND BEST MATCH
# =========================================================

def find_best_match(user_input, index):
    """
    Find best matching prepared row.
    """

    cleaned_input = clean_text(user_input)
    memory["last_match"] = {
        "input": cleaned_input,
        "method": "none",
        "score": 0,
        "second_score": 0,
        "candidates": 0,
        "matched_question": None,
    }

    if not cleaned_input:
        return None, 0

    direct_row, direct_score = direct_intent_match(cleaned_input, index)

    if direct_row:
        memory["last_match"] = {
            "input": cleaned_input,
            "method": "direct intent",
            "score": direct_score,
            "second_score": 0,
            "candidates": 1,
            "matched_question": direct_row["question"],
        }
        return direct_row, direct_score

    # Follow-ups are usually short. Resolve them before the full scan.
    for followup in memory["allowed_followups"]:
        if cleaned_input == followup:
            row = index["followup_lookup"].get(followup)
            memory["last_match"] = {
                "input": cleaned_input,
                "method": "follow-up exact",
                "score": 100,
                "second_score": 0,
                "candidates": len(memory["allowed_followups"]),
                "matched_question": row["question"] if row else None,
            }
            return index["followup_lookup"].get(followup), 100

        if fuzz.partial_ratio(cleaned_input, followup) >= 75:
            row = index["followup_lookup"].get(followup)
            memory["last_match"] = {
                "input": cleaned_input,
                "method": "follow-up fuzzy",
                "score": 100,
                "second_score": 0,
                "candidates": len(memory["allowed_followups"]),
                "matched_question": row["question"] if row else None,
            }
            return row, 100

    tokens = query_tokens(cleaned_input)

    if tokens:
        seen = set()
        candidates = []

        for token in tokens:
            for row in index["token_lookup"].get(token, ()):
                row_id = id(row)

                if row_id not in seen:
                    seen.add(row_id)
                    candidates.append(row)

        candidates = tuple(candidates) or index["rows"]
    else:
        candidates = index["rows"]

    best_row = None
    best_score = 0
    second_score = 0

    for prepared_row in candidates:
        score = calculate_score(cleaned_input, prepared_row)
        overlap_count = len(tokens & prepared_row["tokens"])

        if len(tokens) >= 3 and overlap_count < 2:
            score *= 0.55

        score = apply_context_boost(score, prepared_row, cleaned_input)

        if score > best_score:
            second_score = best_score
            best_score = score
            best_row = prepared_row

            if best_score >= 100:
                break
        elif score > second_score:
            second_score = score

    if not is_confident_match(cleaned_input, best_score, second_score):
        memory["last_match"] = {
            "input": cleaned_input,
            "method": "rejected fuzzy",
            "score": best_score,
            "second_score": second_score,
            "candidates": len(candidates),
            "matched_question": best_row["question"] if best_row else None,
        }
        return None, best_score

    memory["last_match"] = {
        "input": cleaned_input,
        "method": "ranked fuzzy",
        "score": best_score,
        "second_score": second_score,
        "candidates": len(candidates),
        "matched_question": best_row["question"] if best_row else None,
    }

    return best_row, best_score


# =========================================================
# RESPONSE AND MEMORY
# =========================================================

def generate_response(prepared_row):
    responses = prepared_row["responses"] if prepared_row else ()

    if not responses:
        return "Oops, looks like I forgot that one."

    choices = [
        response
        for response in responses
        if response != memory["last_response"]
    ] or list(responses)

    response = format_response(random.choice(choices))
    memory["last_response"] = response

    return response


def format_response(response):
    response = clean_display_text(response)
    mode = memory["answer_mode"]

    if mode == "short":
        sentences = re.split(r"(?<=[.!?])\s+", response)
        return sentences[0] if sentences else response

    if mode == "professional":
        return response.replace("I'm", "I am").replace("I'll", "I will")

    if mode == "recruiter":
        sentences = re.split(r"(?<=[.!?])\s+", response)
        concise = " ".join(sentences[:2]) if sentences else response
        return "Recruiter-ready answer: " + concise

    if mode == "friendly":
        return f"Sure. {response}"

    if mode == "interview":
        return f"In an interview, I would say: {response}"

    return response


def update_memory(prepared_row):
    memory["last_topic"] = prepared_row["question"]
    memory["conversation_depth"] += 1
    memory["allowed_followups"] = prepared_row["followups"]
    memory["last_row"] = prepared_row


def greeting_reply():
    return (
        "Hi, I am Asfaan's portfolio chatbot. "
        "You can ask about his skills, projects, education, strengths, salary, or relocation."
    )


def portfolio_summary_reply(cleaned_input, state):
    tokens = query_tokens(cleaned_input)
    contact = state["contact"]
    project_detail_tokens = {
        "stack",
        "tech",
        "technology",
        "technologies",
        "tools",
        "used",
        "role",
        "contribution",
        "responsibility",
        "responsibilities",
        "impact",
        "result",
        "outcome",
        "challenge",
        "challenges",
        "database",
        "scale",
        "search",
        "feature",
        "features",
    }

    if "vconnect" in cleaned_input or "v connect" in cleaned_input:
        if tokens & project_detail_tokens:
            return None

        return (
            "VConnect is Asfaan's rural analytics dashboard project. It processed data for about 650,000 Indian "
            "villages and made that information easier to search, analyze, and understand. The project used "
            "Python/Pandas for data processing, MongoDB for flexible storage, and an AI-driven search layer to "
            "improve rural data retrieval speed by around 45%."
        )

    if "logisense" in cleaned_input or "logisense 360" in cleaned_input:
        if tokens & project_detail_tokens:
            return None

        return (
            "LogiSense 360 is a subscription billing and management platform. It helps manage customer plans, "
            "usage tracking, invoice generation, payment workflows, and service operations. Asfaan contributed "
            "to development tasks, backend workflow support, debugging, and feature stability."
        )

    if ("nurse" in tokens or "triage" in tokens or "vapi" in tokens) and tokens & {"ai", "assistant", "project", "voice"}:
        return (
            "The AI Nurse Triage Voice Assistant is a healthcare automation project built with Flask and Vapi.ai. "
            "It handled initial patient intake through voice, used NLP intent classification for symptoms, and "
            "helped reduce manual symptom collection time by around 35% while escalating serious cases to nurses."
        )

    if cleaned_input in {"skill", "skills", "technical skill", "technical skills", "tech stack", "stack"}:
        return (
            "Asfaan's core skills are Python, Flask, SQL, MongoDB, Pandas, REST APIs, data processing, "
            "AI voice assistant workflows, NLP intent classification, Git/GitHub, and DSA practice. "
            "His strongest areas are backend development, practical AI integration, and data-heavy projects."
        )

    if cleaned_input in {"project", "projects", "portfolio", "portfolio projects"}:
        return (
            "Asfaan's main projects include: AI Nurse Triage Voice Assistant, which reduced patient intake "
            "effort using voice automation; VConnect, a rural analytics platform for 650,000 Indian village "
            "records; and LogiSense 360, a subscription billing and management system. Ask 'deep dive project' "
            "for details."
        )

    if "github" in tokens and tokens & {"project", "projects", "collaboration", "collaborative", "team"}:
        return (
            "Asfaan uses GitHub to manage project code, track changes, collaborate through commits, and keep "
            "work organized across features and fixes. In project work, GitHub helps show clean version history, "
            "team collaboration, and practical development discipline."
        )

    if cleaned_input in {"education", "study", "degree", "college", "qualification"}:
        return (
            "Asfaan is an MCA student at MEASI Institute of Information Technology. His academic path is focused "
            "on computer applications, software development, databases, and practical AI/data projects."
        )

    if cleaned_input in {"experience", "work experience", "internship", "internships"}:
        return (
            "Asfaan worked as a Software and Data Intern at Smaart Healthcare, where he helped build an AI nurse "
            "assistant, reduced manual patient intake effort, worked on backend/data workflows, and contributed "
            "to real healthcare automation use cases."
        )

    if cleaned_input in {"summary", "profile", "portfolio summary"}:
        return (
            "Mohamed Sathak Asfaan is an MCA student and developer from Chennai focused on Python, Flask, AI, "
            "data processing, and backend systems. His portfolio is strongest around practical AI assistants, "
            "rural data analytics, and workflow automation."
        )

    if tokens & {"email", "phone", "contact", "linkedin"}:
        details = []

        if contact["phone"]:
            details.append(f"Phone: {contact['phone']}")

        if contact["email"]:
            details.append(f"Email: {contact['email']}")

        if contact["linkedin"]:
            details.append(f"LinkedIn: {contact['linkedin']}")

        if contact["github"]:
            details.append(f"GitHub: {contact['github']}")

        if details:
            return " | ".join(details)

    return None


def help_reply(index):
    topics = ", ".join(index["topics"][:8]) if index["topics"] else "skills, projects, education"

    return (
        "Try asking things like: 'Who are you?', 'Tell me about your projects', "
        "'What are your skills?', 'Are you open to relocation?', or 'Expected salary?'. "
        f"Available topic areas include: {topics}. "
        "Commands: mode short, mode interview, mode friendly, mode professional, "
        "mode recruiter, show project questions, practice hr, mock interview, "
        "train, star coach, intro 30, deep dive project, quality check, health, "
        "resume gaps, source on, export transcript, debug on, voice on, reload."
    )


def debug_reply():
    info = memory["last_match"] or {}

    return (
        f"Method: {info.get('method', 'none')} | "
        f"Score: {round(info.get('score', 0), 2)} | "
        f"Second: {round(info.get('second_score', 0), 2)} | "
        f"Candidates checked: {info.get('candidates', 0)} | "
        f"Matched: {info.get('matched_question') or 'none'} | "
        f"Source: {info.get('source') or 'none'}"
    )


def set_answer_mode(cleaned_input):
    parts = cleaned_input.split()

    if len(parts) < 2 or parts[0] != "mode":
        return None

    requested = MODE_ALIASES.get(parts[1], parts[1])

    if requested not in ANSWER_MODES:
        return (
            "I do not know that mode yet. Available modes are: "
            + ", ".join(sorted(ANSWER_MODES))
            + "."
        )

    memory["answer_mode"] = requested
    return f"Answer mode changed to {requested}."


def find_category_rows(index, cleaned_input):
    tokens = query_tokens(cleaned_input)
    rows = []
    alias_targets = {
        clean_text(target)
        for alias, target in CATEGORY_ALIASES.items()
        if clean_text(alias) in cleaned_input
    }

    for row in index["rows"]:
        category_tokens = query_tokens(row["category"])
        question_tokens = query_tokens(row["question"])

        if row["category"] in alias_targets:
            rows.append(row)
        elif tokens & category_tokens or tokens & question_tokens:
            rows.append(row)

    return tuple(rows)


def show_questions_reply(index, cleaned_input):
    rows = find_category_rows(index, cleaned_input)

    if not rows:
        return "I could not find that question category. Try: show project questions, show hr questions, or show skills questions."

    questions = [question_label(row) for row in rows[:10]]
    return "Here are some matching questions: " + " | ".join(questions)


def ask_category_question(index, cleaned_input):
    rows = find_category_rows(index, cleaned_input)

    if not rows:
        return "I could not find that category. Try asking for HR, projects, skills, or behavioral."

    row = random.choice(rows)
    memory["last_row"] = row
    memory["last_topic"] = row["question"]
    return "Practice question: " + question_label(row)


def start_practice(index, cleaned_input):
    rows = find_category_rows(index, cleaned_input)

    if not rows:
        rows = index["rows"]

    rows = list(rows)
    random.shuffle(rows)
    memory["practice"] = {
        "rows": rows,
        "position": 0,
        "current": rows[0],
    }

    return "Practice mode started. " + ask_current_practice_question()


def ask_current_practice_question():
    practice = memory["practice"]
    row = practice["current"]
    return "Question: " + question_label(row)


def handle_practice_answer(user_input):
    practice = memory["practice"]
    row = practice["current"]
    score, tips = score_answer(user_input, row)

    if score < 5:
        feedback = f"Score: {score}/10. Try improving this by: {', '.join(tips)}."
    elif score < 8:
        feedback = f"Score: {score}/10. Decent answer. Improve by: {', '.join(tips)}."
    else:
        feedback = f"Score: {score}/10. Strong answer. Keep the structure clear and concise."

    practice["position"] += 1

    if practice["position"] >= len(practice["rows"]):
        memory["practice"] = None
        return feedback + " Practice complete."

    practice["current"] = practice["rows"][practice["position"]]
    return feedback + " Next " + ask_current_practice_question()


def score_answer(answer, row):
    cleaned = clean_text(answer)
    tokens = query_tokens(cleaned)
    expected_tokens = row["tokens"] or query_tokens(" ".join(row["keywords"]))
    overlap = len(tokens & expected_tokens)
    length_score = min(len(tokens) / 18 * 4, 4)
    keyword_score = min(overlap / 5 * 4, 4)
    example_score = 2 if tokens & {"example", "project", "built", "created", "implemented", "reduced", "improved"} else 0
    score = round(min(length_score + keyword_score + example_score, 10), 1)

    tips = []

    if len(tokens) < 12:
        tips.append("add more detail")

    if overlap < 3:
        tips.append("use more role-specific keywords")

    if example_score == 0:
        tips.append("include a concrete example or measurable result")

    return score, tips or ["good structure and relevance"]


def start_mock_interview(index, cleaned_input):
    rows = list(find_category_rows(index, cleaned_input) or index["rows"])
    random.shuffle(rows)
    rows = rows[:10]

    memory["mock"] = {
        "rows": rows,
        "position": 0,
        "scores": [],
        "current": rows[0],
    }

    return "Mock interview started. " + ask_current_mock_question()


def ask_current_mock_question():
    return "Mock question: " + question_label(memory["mock"]["current"])


def handle_mock_answer(user_input):
    mock = memory["mock"]
    row = mock["current"]
    score, tips = score_answer(user_input, row)
    mock["scores"].append(
        {
            "question": row["question"],
            "score": score,
            "tips": tips,
        }
    )
    mock["position"] += 1

    if mock["position"] >= len(mock["rows"]):
        scores = mock["scores"]
        avg = round(sum(item["score"] for item in scores) / len(scores), 1)
        weak = sorted(scores, key=lambda item: item["score"])[:3]
        memory["mock"] = None
        weak_text = " | ".join(
            f"{question_label({'question': item['question']})} improve: {', '.join(item['tips'])}"
            for item in weak
        )
        return f"Mock interview complete. Final score: {avg}/10. Focus areas: {weak_text}"

    mock["current"] = mock["rows"][mock["position"]]
    return f"Score: {score}/10. Feedback: {', '.join(tips)}. Next {ask_current_mock_question()}"


def star_coach_reply():
    row = memory["last_row"]
    question = question_label(row) if row else "your behavioral answer"

    return (
        f"STAR coach for {question} "
        "Situation: set the context in one line. "
        "Task: explain your responsibility. "
        "Action: describe exactly what you did. "
        "Result: end with a measurable or clear outcome. "
        "Keep it under 90 seconds and use one real example."
    )


def self_intro_reply(cleaned_input):
    if "technical" in cleaned_input:
        return (
            "Technical intro: I am Mohamed Sathak Asfaan, an MCA student focused on Python, Flask, data, "
            "and AI systems. I have built practical projects like an AI nurse assistant and village analytics "
            "tools, where I worked on backend logic, data processing, and real-world workflow improvement."
        )

    if "fresher" in cleaned_input:
        return (
            "Fresher intro: I am Mohamed Sathak Asfaan, an MCA student with strong interest in software "
            "development, AI, and data-driven applications. I enjoy building practical tools, learning quickly, "
            "and applying my skills to solve useful problems."
        )

    if "hr" in cleaned_input:
        return (
            "HR intro: I am Asfaan, an MCA student from Chennai who enjoys building useful technology. "
            "I am comfortable learning fast, working with teams, and taking ownership of tasks from problem "
            "understanding to delivery."
        )

    if "60" in cleaned_input:
        return (
            "60-second intro: I am Mohamed Sathak Asfaan, an MCA student and developer interested in AI, "
            "backend development, and data analytics. During my internship, I worked on an AI nurse assistant "
            "that helped reduce manual patient intake effort. I also built projects such as VConnect, focused "
            "on making rural data easier to access and analyze. I like solving practical problems, improving "
            "systems step by step, and learning from real usage. I am looking for a role where I can contribute, "
            "grow technically, and build software that has visible impact."
        )

    return (
        "30-second intro: I am Mohamed Sathak Asfaan, an MCA student focused on AI, backend development, "
        "and data-driven applications. I have built practical projects including an AI nurse assistant and "
        "rural analytics tools, and I enjoy turning real problems into working software."
    )


def project_deep_dive_reply(index, cleaned_input):
    rows = [
        row
        for row in find_category_rows(index, "project intern " + cleaned_input)
        if row["category"] == "projects intern"
    ]

    if not rows:
        rows = [row for row in index["rows"] if row["category"] == "projects intern"]

    selected = rows[:5]
    points = []

    for row in selected:
        response = row["responses"][0] if row["responses"] else ""
        points.append(f"{question_label(row)} {clean_display_text(response)[:180]}")

    return "Project deep dive: " + " | ".join(points)


def export_transcript():
    TRANSCRIPT_DIR.mkdir(exist_ok=True)
    path = TRANSCRIPT_DIR / f"chat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    with path.open("w", encoding="utf-8") as file:
        for item in memory["transcript"]:
            if "user" in item:
                file.write(f"[{item['time']}] You: {item['user']}\n")
            if "bot" in item:
                file.write(f"[{item['time']}] Bot: {item['bot']}\n")

    return f"Transcript exported to {path}"


def health_reply(state):
    missed_count = 0

    if MISSED_QUESTIONS_FILE.exists():
        missed_count = len(MISSED_QUESTIONS_FILE.read_text(encoding="utf-8").splitlines())

    return (
        f"Health: rows={len(state['index']['rows'])}, "
        f"resume_chunks={len(state['resume_index']['chunks'])}, "
        f"knowledge_projects={len(state.get('knowledge_base', {}).get('projects', {}))}, "
        f"similarity_docs={len(state.get('similarity_docs', ()))}, "
        f"rapidfuzz={RAPIDFUZZ_AVAILABLE}, pdf_reader={PDF_READER_AVAILABLE}, "
        f"voice_output={TTS_AVAILABLE}, voice_input={SPEECH_RECOGNITION_AVAILABLE}, "
        f"excel_update={OPENPYXL_AVAILABLE}, missed_questions={missed_count}, "
        f"mode={memory['answer_mode']}."
    )


def quality_check_reply(state):
    rows = state["index"]["rows"]
    empty_answers = sum(1 for row in rows if not row["responses"])
    weak_keywords = sum(1 for row in rows if len(row["tokens"]) < 4)
    questions = [row["question"] for row in rows]
    duplicates = len(questions) - len(set(questions))
    no_followups = sum(1 for row in rows if not row["followups"])

    return (
        f"Quality check: empty_answers={empty_answers}, duplicate_questions={duplicates}, "
        f"weak_keyword_rows={weak_keywords}, rows_without_followups={no_followups}. "
        "Best next cleanup: add stronger casual phrases and follow-ups for rows with weak keywords."
    )


def suggestions_reply(state):
    project_names = [
        project["name"]
        for project in state.get("knowledge_base", {}).get("projects", {}).values()
    ]
    rows = state["index"]["rows"]
    sample = random.sample(rows, min(3, len(rows)))
    labels = " | ".join(project_names[:3] + [question_label(row) for row in sample[:2]])
    return "I am not fully sure. Did you mean one of these: " + labels


def source_labeled(response, source):
    if not memory["show_source"]:
        return response

    return f"[Source: {source}] {response}"


def resume_gap_reply(state):
    resume_tokens = set()

    for chunk in state["resume_index"]["chunks"]:
        resume_tokens |= chunk["tokens"]

    gaps = []

    for row in state["index"]["rows"]:
        row_tokens = row["tokens"]

        if not row_tokens:
            gaps.append(row)
            continue

        overlap = len(row_tokens & resume_tokens) / max(len(row_tokens), 1)

        if overlap < 0.12:
            gaps.append(row)

    labels = " | ".join(question_label(row) for row in gaps[:10])

    if not gaps:
        return "Resume gap detector: the resume covers the current workbook topics reasonably well."

    return (
        f"Resume gap detector: {len(gaps)} workbook questions have weak resume evidence. "
        f"Examples: {labels}"
    )


def append_training_row(question, answer, category="User Trained"):
    if not OPENPYXL_AVAILABLE:
        return "Excel updater needs openpyxl installed."

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    next_row = sheet.max_row + 1
    row_values = {header: "" for header in headers}
    row_values["Q#"] = next_row
    row_values["Category"] = category
    row_values["Question"] = question
    row_values["Casual 1"] = question
    row_values["Fuzzy Logic 1"] = question
    row_values["Calm Response"] = answer
    row_values["Friendly Response"] = answer

    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=next_row, column=col_index, value=row_values.get(header, ""))

    workbook.save(EXCEL_FILE)
    clean_text.cache_clear()
    return f"Learned and added to Excel: {question}"


def start_trainer():
    last_question = None

    if MISSED_QUESTIONS_FILE.exists():
        lines = MISSED_QUESTIONS_FILE.read_text(encoding="utf-8").splitlines()
        if lines:
            last_question = lines[-1].split("\t")[-1]

    memory["trainer"] = {
        "step": "answer" if last_question else "question",
        "question": last_question,
    }

    if last_question:
        return f"Trainer mode: I found the last missed question: '{last_question}'. Type the answer I should learn."

    return "Trainer mode: type the question I should learn."


def handle_trainer_input(user_input, state):
    trainer = memory["trainer"]

    if clean_text(user_input) in {"cancel", "stop trainer", "end trainer"}:
        memory["trainer"] = None
        return "Trainer mode cancelled."

    if trainer["step"] == "question":
        trainer["question"] = clean_display_text(user_input)
        trainer["step"] = "answer"
        return "Got the question. Now type the answer I should give."

    question = trainer["question"]
    answer = clean_display_text(user_input)
    result = append_training_row(question, answer)
    memory["trainer"] = None
    state.update(load_bot_state())
    return result


def learn_one_line(cleaned_input, user_input, state):
    raw = user_input.strip()
    payload = raw.split(":", 1)[1].strip() if ":" in raw else raw

    if "=>" not in payload:
        return "Use this format: learn: question => answer"

    question, answer = [part.strip() for part in payload.split("=>", 1)]

    if not question or not answer:
        return "Both question and answer are needed."

    result = append_training_row(question, answer)
    state.update(load_bot_state())
    return result


# =========================================================
# FALLBACK RESPONSES
# =========================================================

fallback_responses = (
    "Hmm, I don't think I learned that one yet.",
    "Okay, that's a new question.",
    "I might need a software update for that one.",
    "You just discovered one of my weak spots.",
    "That topic isn't inside my portfolio brain yet.",
    "Interesting one, but I don't have data for that yet.",
)


def fallback_reply():
    return random.choice(fallback_responses)


def log_missed_question(user_input, score):
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"{timestamp}\tScore={round(score, 2)}\t{clean_display_text(user_input)}\n"
        with MISSED_QUESTIONS_FILE.open("a", encoding="utf-8") as file:
            file.write(line)
    except Exception:
        pass


def resume_reply(chunk):
    return format_response(f"From the resume: {chunk['text']}")


def question_label(row):
    return str(row["question"]).strip().rstrip(".?!").capitalize() + "?"


def load_bot_state():
    df = load_dataset(EXCEL_FILE)
    index = prepare_dataset(df)
    resume_text = load_resume_text(RESUME_FILE)
    resume_index = prepare_resume_index(resume_text)
    knowledge_base = load_knowledge_base(KNOWLEDGE_BASE_FILE)
    state = {
        "index": index,
        "resume_index": resume_index,
        "contact": extract_contact_info(resume_text),
        "knowledge_base": knowledge_base,
    }
    state["similarity_docs"] = build_similarity_documents(state)

    return state


def process_user_input(user_input, state):
    remember_user_message(user_input)
    cleaned_input = clean_text(user_input)

    if not cleaned_input:
        return finish_response("", False)

    index = state["index"]
    resume_index = state["resume_index"]

    if cleaned_input in EXIT_COMMANDS:
        return finish_response("Catch you later.", True)

    if memory["trainer"]:
        return finish_response(handle_trainer_input(user_input, state), False)

    if memory["mock"]:
        if cleaned_input in {"stop mock", "end mock", "quit mock"}:
            memory["mock"] = None
            return finish_response("Mock interview stopped.", False)

        return finish_response(handle_mock_answer(user_input), False)

    if memory["practice"]:
        if cleaned_input in PRACTICE_STOP_COMMANDS:
            memory["practice"] = None
            return finish_response("Practice mode stopped.", False)

        return finish_response(handle_practice_answer(user_input), False)

    if cleaned_input in GREETING_PATTERNS:
        return finish_response(greeting_reply(), False)

    if cleaned_input in HELP_COMMANDS:
        return finish_response(help_reply(index), False)

    if cleaned_input.startswith("diagnose "):
        query = cleaned_input.replace("diagnose", "", 1).strip()
        answer, score, source = find_similarity_answer(query, state)
        kb_answer, kb_source = knowledge_answer(query, state)
        return finish_response(
            "Diagnosis: "
            f"intent_source={kb_source or 'none'} | "
            f"similarity_source={source or 'none'} | "
            f"similarity_score={round(score, 3)} | "
            f"preview={(kb_answer or answer or 'no confident answer')[:180]}",
            False,
        )

    portfolio_reply = portfolio_summary_reply(cleaned_input, state)

    if portfolio_reply:
        return finish_response(source_labeled(portfolio_reply, "portfolio summary"), False)

    if cleaned_input == "health":
        return finish_response(health_reply(state), False)

    if cleaned_input in {"quality check", "data quality", "scan excel", "check excel"}:
        return finish_response(quality_check_reply(state), False)

    if cleaned_input in {"resume gaps", "resume gap", "gap detector", "resume gap detector"}:
        return finish_response(resume_gap_reply(state), False)

    if cleaned_input in {"export transcript", "save transcript", "transcript export"}:
        return finish_response(export_transcript(), False)

    if cleaned_input in {"suggest", "suggestions", "did you mean"}:
        return finish_response(suggestions_reply(state), False)

    if cleaned_input in {"train", "trainer", "trainer mode"}:
        return finish_response(start_trainer(), False)

    if cleaned_input.startswith("learn"):
        return finish_response(learn_one_line(cleaned_input, user_input, state), False)

    if cleaned_input.startswith("mock"):
        return finish_response(start_mock_interview(index, cleaned_input), False)

    if cleaned_input in {"star", "star coach", "star method", "coach star"}:
        return finish_response(star_coach_reply(), False)

    if (
        cleaned_input.startswith("intro")
        or "self introduction" in cleaned_input
        or "fresher intro" in cleaned_input
        or "technical intro" in cleaned_input
        or "hr intro" in cleaned_input
    ):
        return finish_response(self_intro_reply(cleaned_input), False)

    if cleaned_input.startswith("deep dive") or cleaned_input.startswith("project deep dive"):
        return finish_response(project_deep_dive_reply(index, cleaned_input), False)

    kb_response, kb_source = knowledge_answer(cleaned_input, state)

    if kb_response:
        memory["last_match"] = {
            "input": cleaned_input,
            "method": "intent classifier",
            "score": 100,
            "second_score": 0,
            "candidates": len(state.get("knowledge_base", {}).get("projects", {})),
            "matched_question": kb_source,
            "source": kb_source,
        }
        return finish_response(source_labeled(kb_response, kb_source), False)

    contact_tokens = query_tokens(cleaned_input)
    wants_contact = bool(contact_tokens & {"phone", "number", "email", "mail", "linkedin", "contact"})
    wants_contact = wants_contact or bool(
        "github" in contact_tokens
        and contact_tokens & {"profile", "account", "link", "url"}
    )

    if wants_contact:
        contact = state["contact"]
        details = []

        if contact["phone"]:
            details.append(f"Phone: {contact['phone']}")

        if contact["email"]:
            details.append(f"Email: {contact['email']}")

        if contact["linkedin"]:
            details.append(f"LinkedIn: {contact['linkedin']}")

        if contact["github"]:
            details.append(f"GitHub: {contact['github']}")

        if details:
            return finish_response(" | ".join(details), False)

    mode_reply = set_answer_mode(cleaned_input)

    if mode_reply:
        return finish_response(mode_reply, False)

    if cleaned_input in {"debug", "why this", "why did you answer that"}:
        return finish_response(debug_reply(), False)

    if cleaned_input == "voice on":
        if TTS_AVAILABLE:
            memory["speak"] = True
            return finish_response("Voice output is on.", False)

        return finish_response("Voice output needs pyttsx3 installed. The chatbot will keep using text for now.", False)

    if cleaned_input == "voice off":
        memory["speak"] = False
        return finish_response("Voice output is off.", False)

    if cleaned_input == "source on":
        memory["show_source"] = True
        return finish_response("Source labels are on.", False)

    if cleaned_input == "source off":
        memory["show_source"] = False
        return finish_response("Source labels are off.", False)

    if cleaned_input in {"listen", "voice input"}:
        spoken_text, error = listen_for_voice()

        if error:
            return finish_response(error, False)

        response, should_exit = process_user_input(spoken_text, state)
        return finish_response(f"I heard: {spoken_text}. {response}", should_exit)

    if cleaned_input.startswith("show ") and "question" in cleaned_input:
        return finish_response(show_questions_reply(index, cleaned_input), False)

    if cleaned_input.startswith("ask "):
        return finish_response(ask_category_question(index, cleaned_input), False)

    if cleaned_input.startswith("practice"):
        return finish_response(start_practice(index, cleaned_input), False)

    if cleaned_input == "debug on":
        globals()["DEBUG_SCORE"] = True
        return finish_response("Debug scores are now visible.", False)

    if cleaned_input == "debug off":
        globals()["DEBUG_SCORE"] = False
        return finish_response("Debug scores are now hidden.", False)

    if cleaned_input == "reload":
        state.update(load_bot_state())
        return finish_response(
            f"Reloaded {len(state['index']['rows'])} rows from the workbook and "
            f"{len(state['resume_index']['chunks'])} resume chunks."
        , False)

    best_row, best_score = find_best_match(user_input, index)

    if DEBUG_SCORE:
        print(f"DEBUG: {debug_reply()}")

    if best_row and best_score >= FUZZY_THRESHOLD:
        update_memory(best_row)
        response = generate_response(best_row)

        if validate_answer(cleaned_input, response, "excel"):
            return finish_response(source_labeled(response, memory["last_match"]["method"]), False)

    similarity_response, similarity_score, similarity_source = find_similarity_answer(cleaned_input, state)

    if similarity_response:
        memory["last_match"] = {
            "input": cleaned_input,
            "method": "local similarity",
            "score": similarity_score,
            "second_score": 0,
            "candidates": len(state.get("similarity_docs", ())),
            "matched_question": similarity_source,
            "source": similarity_source,
        }

        if validate_answer(cleaned_input, similarity_response, similarity_source):
            return finish_response(source_labeled(format_response(similarity_response), similarity_source), False)

    resume_chunk, resume_score = find_resume_match(user_input, resume_index)

    if resume_chunk:
        memory["last_match"] = {
            "input": cleaned_input,
            "method": "resume fallback",
            "score": resume_score,
            "second_score": 0,
            "candidates": len(resume_index["chunks"]),
            "matched_question": "resume",
        }
        return finish_response(source_labeled(resume_reply(resume_chunk), "resume"), False)

    log_missed_question(user_input, best_score)
    return finish_response(fallback_reply() + " " + suggestions_reply(state), False)


# =========================================================
# TERMINAL CHAT LOOP
# =========================================================

def start_chatbot():
    print("\n=================================================")
    print("Mohamed Sathak Asfaan Portfolio Chatbot")
    print("=================================================")
    print("Type 'exit' anytime to quit.\n")

    try:
        state = load_bot_state()
        index = state["index"]
        resume_index = state["resume_index"]

        print("Dataset loaded successfully")
        print(f"Rows loaded: {len(index['rows'])}\n")
        if resume_index["chunks"]:
            print(f"Resume knowledge loaded: {len(resume_index['chunks'])} chunks\n")

    except Exception as e:
        print(f"Failed to load dataset:\n{e}")
        return

    while True:
        user_input = input("You: ")

        if not user_input.strip():
            continue

        try:
            response, should_exit = process_user_input(user_input, state)
        except Exception as e:
            response, should_exit = f"Something went wrong while answering: {e}", False

        if response:
            print_bot(response)

        if should_exit:
            break


def start_gui():
    try:
        import tkinter as tk
        from tkinter import scrolledtext
    except Exception as e:
        print(f"Desktop GUI is not available on this Python install: {e}")
        return

    try:
        state = load_bot_state()
    except Exception as e:
        print(f"Failed to load chatbot data: {e}")
        return

    root = tk.Tk()
    root.title("Asfaan Portfolio Chatbot")
    root.geometry("820x620")

    chat = scrolledtext.ScrolledText(root, wrap=tk.WORD, state="disabled", padx=12, pady=12)
    chat.pack(fill=tk.BOTH, expand=True)

    entry = tk.Entry(root)
    entry.pack(fill=tk.X, padx=10, pady=10)

    def add_message(sender, message):
        chat.configure(state="normal")
        chat.insert(tk.END, f"{sender}: {clean_display_text(message)}\n\n")
        chat.configure(state="disabled")
        chat.see(tk.END)

    def send_message(event=None):
        user_input = entry.get().strip()

        if not user_input:
            return

        entry.delete(0, tk.END)
        add_message("You", user_input)

        try:
            response, should_exit = process_user_input(user_input, state)
        except Exception as e:
            response, should_exit = f"Something went wrong while answering: {e}", False

        add_message("Bot", response)

        if should_exit:
            root.after(400, root.destroy)

    entry.bind("<Return>", send_message)

    send_button = tk.Button(root, text="Send", command=send_message)
    send_button.pack(pady=(0, 10))

    add_message(
        "Bot",
        "Hi, I am ready. Ask about skills, projects, education, relocation, salary, or type help.",
    )

    entry.focus_set()
    root.mainloop()


# =========================================================
# WEB API FOR PORTFOLIO INTEGRATION
# =========================================================

def get_web_state():
    global web_state

    if web_state is None:
        web_state = load_bot_state()

    return web_state


def create_web_app():
    if not FLASK_AVAILABLE:
        raise RuntimeError("Flask is not installed. Run: pip install -r requirements.txt")

    api = Flask(__name__)

    @api.after_request
    def add_cors_headers(response):
        response.headers["Access-Control-Allow-Origin"] = os.environ.get("CHATBOT_ALLOWED_ORIGIN", "*")
        response.headers["Access-Control-Allow-Headers"] = "Content-Type"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return response

    @api.get("/api/health")
    def api_health():
        state = get_web_state()
        return jsonify(
            {
                "ok": True,
                "openrouter_configured": openrouter_is_configured(),
                "openrouter_model": OPENROUTER_MODEL if openrouter_is_configured() else None,
                "rows": len(state["index"]["rows"]),
                "resume_chunks": len(state["resume_index"]["chunks"]),
                "projects": len(state.get("knowledge_base", {}).get("projects", {})),
            }
        )

    @api.post("/api/chat")
    def api_chat():
        payload = request.get_json(silent=True) or {}
        message = str(payload.get("message", "")).strip()

        if not message:
            return jsonify({"reply": "Ask me something about Asfaan's projects, skills, education, or experience."})

        try:
            local_response, should_exit = process_user_input(message, get_web_state())
            openrouter_response = openrouter_generate_reply(message, local_response)
            return jsonify(
                {
                    "reply": openrouter_response or local_response,
                    "openrouter_configured": openrouter_is_configured(),
                    "openrouter_error": None if openrouter_response else last_openrouter_error,
                    "openrouter_requested_model": OPENROUTER_MODEL if openrouter_is_configured() else None,
                    "openrouter_selected_model": last_openrouter_selected_model if openrouter_response else None,
                    "source": "openrouter" if openrouter_response else "local",
                    "should_exit": should_exit,
                }
            )
        except Exception as e:
            return jsonify({"reply": f"Something went wrong while answering: {e}", "should_exit": False}), 500

    @api.post("/api/reload")
    def api_reload():
        global web_state
        web_state = load_bot_state()
        return jsonify({"ok": True, "message": "Chatbot data reloaded."})

    return api


def start_web_api():
    host = os.environ.get("CHATBOT_HOST", "127.0.0.1")
    port = int(os.environ.get("CHATBOT_PORT", "5000"))
    create_web_app().run(host=host, port=port, debug=False)


app = create_web_app() if FLASK_AVAILABLE else None


# =========================================================
# MAIN
# =========================================================

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1].lower() in {"api", "--api", "server", "web"}:
        start_web_api()
    elif len(sys.argv) > 1 and sys.argv[1].lower() in {"gui", "--gui", "desktop"}:
        start_gui()
    else:
        start_chatbot()
